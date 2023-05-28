"""Primary analysis script."""

__author__ = "Matt Shields"
__copyright__ = "Copyright 2022, National Renewable Energy Laboratory"
__maintainer__ = "Matt Shields"
__email__ = "matt.shields@nrel.gov"

import openpyxl
import pandas as pd
import datetime as dt
import pprint

from ORBIT import load_config
from ORBIT import ProjectManager
from ORBIT.core.library import initialize_library

import os
west_coast_dir = '/Users/asharma/codes/P_Code/currTests/west_coast_cost_modeling/west_coast_sites'
# west_coast_dir = 'C:/Users/mshields/Documents/Projects/West Coast ports/Analysis/LCOE/west_coast_cost_modeling/west_coast_sites'
os.chdir(west_coast_dir)
write_mode = True

start_dates = ['01/01/2002', '01/01/2005', '01/01/2008', '01/01/2011', '01/01/2014', '01/01/2017', '01/01/2020']

for start_date in start_dates:

    # site problem parameters
    site = 'southern_WA'
    mean_windspeed = 8.31
    depth = 913
    distance_to_landfall = 89.117

    # port problem parameters
    port = 'Grays_Harbor'
    distance_to_site = 89.117 # port distance to site
    sub_assembly_lines = 1
    turbine_assembly_cranes = sub_assembly_lines
    sub_storage = 5
    assembly_storage = 5

    if 'DATA_LIBRARY' in os.environ:
        del os.environ['DATA_LIBRARY']

    # set relative paths (alternatively, set absolute paths)
    custom_library = 'data'
    custom_config  = 'base_setup.yaml'
    custom_weather = 'data/weather/' + site + '_swh_150m.csv'

    if __name__ == '__main__':
        # Point ORBIT to the custom data libraries in the anlaysis repo
        initialize_library(custom_library)

        # Load in the input configuration YAML
        base_config = load_config(custom_config)

        # configuration to be modified
        mod_config = {
            'site': {
            'mean_windspeed': mean_windspeed,
            'distance': distance_to_site,
            'depth': depth,
            'distance_to_landfall': distance_to_landfall
            },

            'port': {
            'name': port,
            'sub_assembly_lines': sub_assembly_lines,
            'sub_storage': sub_storage,
            'turbine_assembly_cranes': turbine_assembly_cranes,
            'assembly_storage': assembly_storage
            },

            'install_phases': {
            'MooringSystemInstallation': start_date,
            'MooredSubInstallation': ('MooringSystemInstallation', 0.4),
            'ArrayCableInstallation': ('MooredSubInstallation', 0.8),
            'ExportCableInstallation': start_date,
            'FloatingSubstationInstallation': ('ExportCableInstallation', 0.25)
            }
        }

        # create run config
        run_config = ProjectManager.merge_dicts(base_config, mod_config)

        # Print out the required information for input config
        phases = ['ArraySystemDesign',
                'ElectricalDesign',
                'SemiSubmersibleDesign',
                'SemiTautMooringSystemDesign',
                'ArrayCableInstallation',
                'ExportCableInstallation',
                'MooringSystemInstallation',
                'FloatingSubstationInstallation',
                'MooredSubInstallation']
        expected_config = ProjectManager.compile_input_dict(phases)
        pp = pprint.PrettyPrinter(indent=4)
        # pp.pprint(expected_config)

        # Initialize and run project
        weather = pd.read_csv(custom_weather, parse_dates=["datetime"]).set_index("datetime")
        project = ProjectManager(run_config, weather=weather)
        project.run(include_onshore_construction=False)

        # Print some output results
        print("\n")
        pp.pprint(project.capex_breakdown_per_kw)

        print(f"\nInstallation CapEx: {project.installation_capex/1e6:.0f} M")
        print(f"System CapEx: {project.system_capex/1e6:.0f} M")
        print(f"Turbine CapEx: {project.turbine_capex/1e6:.0f} M")
        print(f"Soft CapEx: {project.soft_capex/1e6:.0f} M")
        print(f"Total CapEx: {project.total_capex/1e6:.0f} M\n")
    
        # print detailed ouptus
        # pp.pprint(project.detailed_outputs)

        # print phase dates
        pp.pprint(project.phase_dates)

        # Should add a method here to report the start/end dates of each phase and maybe plot a Gantt chart or similar
        df = pd.DataFrame(project.actions)
        if write_mode:
            time_str = pd.to_datetime(start_date)
            df.to_excel('scenario_actions/' + site + '_action_' + port + '_' + time_str.strftime('%m_%d_%Y') + '.xlsx', index=False)

        print("\n")
        pp.pprint(project.phase_times)

        print(f"\nTotal Installation Time: {df['time'].iloc[-1]:.0f} h")
        print(f"Total Installation Time: {(df['time'].iloc[-1])/24:.0f} days")

        ## write to excel sheet
        # workbook = openpyxl.load_workbook('/Users/asharma/Downloads/orbit_scenarios.xlsx')
        # worksheet = workbook['Summary']
        # # Iterate through the rows to find the desired matches
        # row_id = 1
        # for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, values_only=True):
        #     if row[0] and row[0].lower() == site.lower():
        #         if row[1] and row[1] == mean_windspeed:
        #             if row[2] and row[2] == depth:
        #                 if row[3] and row[3].lower() == (str(distance_to_site) + ' ' + '(' + port + ')').lower():
        #                     if row[4] and str(distance_to_landfall) in row[4]:
        #                         if row[5] and row[5] == start_date:
        #                             print('\nFOUND FOUND FOUND FOUND FOUND FOUND FOUND\n')
        #                             # Found a row with the desired matches
        #                             # Write values to cells in that row
        #                             worksheet.cell(row=row_id, column=7, value=df['time'].iloc[-1])
        #                             worksheet.cell(row=row_id, column=7).number_format = '0'
        #                             worksheet.cell(row=row_id, column=8, value=(df['time'].iloc[-1])/24)
        #                             worksheet.cell(row=row_id, column=8).number_format = '0'
        #                             worksheet.cell(row=row_id, column=9, value=project.total_capex/1e6)
        #                             worksheet.cell(row=row_id, column=9).number_format = '0'
        #                             break  # Stop iterating after finding the first match
        #     row_id += 1
        # #save
        # workbook.save('/Users/asharma/Downloads/orbit_scenarios.xlsx')